#!/usr/bin/env python3
"""
Register PDF -> Excel digitization tool (R.R. Garden guest register).

A scanned handwritten register is digitized in a loop:
  verify -> (render -> [agent reads images & transcribes] -> save)* -> build

The only step a human/AI must do manually is READ the rendered images and
produce per-page JSON (see TRANSCRIPTION_GUIDE.md). Everything else is scripted.

Commands
  python3 digitize.py verify            show PDF->page mapping + render samples of
                                        logical page 1 so you can confirm group_size
  python3 digitize.py render <P>        render all images needed to transcribe
                                        logical register page P
  python3 digitize.py save <P>          read rows JSON on stdin, save data/pagePP.json
  python3 digitize.py status            which pages are done / what is next
  python3 digitize.py build             rebuild the completed Excel from data/*.json

Page grouping: group_size in config.json controls how many PDF pages form ONE
logical register page (2 = register page spread across left+right PDF pages,
1 = one PDF page per register page, 3 = three, etc.).
Logical page P occupies global PDF pages [ (P-1)*group_size+1 ... P*group_size ].
"""
import sys, os, json, glob, re, copy

CFG = json.load(open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')))

def pdf_files():
    rx = re.compile(CFG['pdf_sort_regex'])
    fs = sorted(glob.glob(CFG['pdf_glob']), key=lambda p: int(rx.search(p).group(1)))
    if not fs:
        sys.exit(f"No PDFs match {CFG['pdf_glob']}")
    return fs

def page_map():
    """Return list of (file, page_index) indexed by global page number (1-based)."""
    import fitz
    out = [None]
    for f in pdf_files():
        for i in range(len(fitz.open(f))):
            out.append((f, i))
    return out

def render_page(gpage, zoom, clip_frac=None, suffix=''):
    import fitz
    pm = page_map()
    if gpage >= len(pm):
        sys.exit(f"global page {gpage} out of range (max {len(pm)-1})")
    f, idx = pm[gpage]
    pg = fitz.open(f)[idx]
    W, H = pg.rect.width, pg.rect.height
    clip = None
    if clip_frac:
        x0, y0, x1, y1 = clip_frac
        clip = fitz.Rect(W*x0, H*y0, W*x1, H*y1)
    pix = pg.get_pixmap(matrix=fitz.Matrix(zoom, zoom), clip=clip)
    os.makedirs(CFG['render_dir'], exist_ok=True)
    out = os.path.join(CFG['render_dir'], f"g{gpage:03d}{suffix}.jpg")
    pix.save(out, jpg_quality=82)
    print(out, pix.width, pix.height)
    return out

def cmd_verify():
    import fitz
    print("PDF -> global page mapping:")
    g = 1
    for f in pdf_files():
        n = len(fitz.open(f))
        print(f"  {os.path.basename(f):45s} {n:3d} pages -> global {g}-{g+n-1}")
        g += n
    total = g - 1
    gs = CFG['group_size']
    print(f"\nTotal {total} PDF pages; group_size={gs} -> {total // gs} logical register pages")
    print("Rendering logical page 1 samples (check them visually):")
    for k in range(gs):
        render_page(1 + k, CFG['zoom_full'], suffix='_verify')
    print("""
CHECK: open the rendered files. For a 2-page spread register:
  - odd page = S.No / Name / Age / Relationship / Nationality / Persons / Address
  - even page = Phone / Coming From / Destination / Purpose / Arrival / Departure / Room
  - the even (right) page usually carries a circled page-number stamp top-right.
If one PDF page shows ALL columns, set group_size=1 in config.json.
If S.No continue across 3 pages before repeating, set group_size=3. Re-run verify.""")

def cmd_render(P):
    gs = CFG['group_size']
    pages = [(P-1)*gs + 1 + k for k in range(gs)]
    for i, g in enumerate(pages):
        render_page(g, CFG['zoom_full'])                       # full page
        if gs == 2 and i == 1:                                 # right page extras
            x1 = CFG['right_page_crop_x1']
            render_page(g, CFG['zoom_crop'], (0, 0.08, x1, 0.56), '_c0')
            render_page(g, CFG['zoom_crop'], (0, 0.52, x1, 1.00), '_c1')
            for si, (a, b) in enumerate([(0.10, 0.42), (0.40, 0.72), (0.70, 1.0)]):
                render_page(g, CFG['zoom_strip'], (0, a, CFG['phone_strip_x1'], b), f'_s{si}')
    print(f"\nLogical page {P} rendered (global pages {pages}). Read the images and")
    print(f"then:  python3 digitize.py save {P} < rows.json   (schema in TRANSCRIPTION_GUIDE.md)")

def cmd_save(P):
    rows = json.load(sys.stdin)
    os.makedirs(CFG['data_dir'], exist_ok=True)
    for r in rows:
        r['page'] = P
    with open(os.path.join(CFG['data_dir'], f'page{P:02d}.json'), 'w') as f:
        json.dump(rows, f, indent=1, ensure_ascii=False)
    print(f'saved page {P}: {len(rows)} rows')
    cmd_status()

def done_pages():
    return sorted(int(re.search(r'page(\d+)', p).group(1))
                  for p in glob.glob(os.path.join(CFG['data_dir'], 'page*.json')))

def cmd_status():
    d = done_pages()
    total = (len(page_map()) - 1) // CFG['group_size']
    nxt = next((p for p in range(1 + CFG['first_record_page_already_done'], total + 1)
                if p not in d), None)
    print(f'pages done: {d}\nnext page to transcribe: {nxt} (of {total})')

def cmd_build():
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.load_workbook(CFG['tracker_xlsx'])
    ws = wb[CFG['sheet']]
    tmpl = [ws.cell(row=CFG['template_row'], column=c) for c in range(1, 16)]
    yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    base = CFG['append_after_row']
    if ws.max_row > base:
        ws.delete_rows(base + 1, ws.max_row - base)
    r = base + 1
    count = 0
    for P in done_pages():
        for row in json.load(open(os.path.join(CFG['data_dir'], f'page{P:02d}.json'))):
            vals = [P, row.get('sno'), row.get('name'), row.get('age'), row.get('rel'),
                    row.get('nat'), row.get('num'), row.get('addr'), row.get('phone'),
                    row.get('from'), row.get('dest'), row.get('purpose'), row.get('arr'),
                    row.get('dep'), row.get('room')]
            for c, (v, t) in enumerate(zip(vals, tmpl), start=1):
                cell = ws.cell(row=r, column=c, value=v if v not in (None, '') else None)
                cell.font = copy.copy(t.font); cell.border = copy.copy(t.border)
                cell.alignment = copy.copy(t.alignment)
                if isinstance(v, str) and 'VERIFY' in v:
                    cell.fill = yellow
            r += 1; count += 1
    # monthly bookings + commission sheet
    title = 'Monthly Bookings & Ansh'
    if title in wb.sheetnames:
        del wb[title]
    s = wb.create_sheet(title)
    s['A1'] = 'Monthly Bookings & Ansh Commission'; s['A1'].font = Font(bold=True, size=14)
    rs = CFG['commission_per_booking_rs']
    s['A2'] = (f'Booking count from Guest Register ArrivalDateTime. Commission = Rs {rs}/booking; '
               'manager remits within 15 days minus commission.')
    s['A2'].font = Font(italic=True, size=9)
    for i, h in enumerate(['Month', 'Bookings', f'Commission (Rs {rs}/booking)'], 1):
        c = s.cell(row=4, column=i, value=h); c.font = Font(bold=True)
    months = [(m, y) for y in (24, 25, 26) for m in range(1, 13)][10:31]  # Nov24..May26
    col = f"'{CFG['sheet']}'!$M$5:$M$10000"
    ri = 5
    for m, y in months:
        s.cell(row=ri, column=1, value=f'{m:02d}/20{y}')
        pats = [f'*/{m}/{y} *', f'*/{m}/{y}', f'*/0{m}/{y} *', f'*/0{m}/{y}'] if m < 10 else [f'*/{m}/{y} *', f'*/{m}/{y}']
        s.cell(row=ri, column=2, value='=' + '+'.join(f'COUNTIF({col},"{p}")' for p in pats))
        s.cell(row=ri, column=3, value=f'=B{ri}*{rs}')
        ri += 1
    s.cell(row=ri, column=1, value='TOTAL').font = Font(bold=True)
    s.cell(row=ri, column=2, value=f'=SUM(B5:B{ri-1})').font = Font(bold=True)
    s.cell(row=ri, column=3, value=f'=SUM(C5:C{ri-1})').font = Font(bold=True)
    os.makedirs(os.path.dirname(CFG['output_xlsx']), exist_ok=True)
    wb.save(CFG['output_xlsx'])
    print(f'appended {count} rows -> {CFG["output_xlsx"]}')

if __name__ == '__main__':
    a = sys.argv[1:]
    if not a: sys.exit(__doc__)
    cmd = a[0]
    if cmd == 'verify': cmd_verify()
    elif cmd == 'render': cmd_render(int(a[1]))
    elif cmd == 'save': cmd_save(int(a[1]))
    elif cmd == 'status': cmd_status()
    elif cmd == 'build': cmd_build()
    else: sys.exit(__doc__)
